# ============================================================
# utils/exporter.py — Exportación a Excel y gestión de historial
# ============================================================

import os
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import DIRECTORIO_OUTPUT, ARCHIVO_HISTORIAL


def exportar_a_excel(termino: str, ofertas: list) -> str:
    """
    Genera un reporte en Excel (.xlsx) altamente estilizado y profesional
    con los resultados reales encontrados, ordenados por precio total.
    """
    if not os.path.exists(DIRECTORIO_OUTPUT):
        os.makedirs(DIRECTORIO_OUTPUT, exist_ok=True)

    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = "".join([c if c.isalnum() else "_" for c in termino])
    archivo_salida = os.path.join(DIRECTORIO_OUTPUT, f"precios_{nombre_limpio}_{fecha_str}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparador de Precios"
    ws.views.sheetView[0].showGridLines = True

    # 1. Título principal
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"REPORTE COMPARATIVO DE PRECIOS — {termino.upper()}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 2. Metadatos
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Desarrollado por Felix Borges (Scraper Python)"
    ws["A2"].font = Font(size=9, italic=True, color="595959")
    ws.row_dimensions[2].height = 18

    # 3. Encabezados de tabla
    headers = ["Pos.", "Tienda", "Producto / Oferta", "Desc. %", "Precio Base", "Envío", "TOTAL (USD)", "Enlace a Tienda"]
    ws.append([]) # Fila 3 vacía
    ws.append(headers) # Fila 4
    ws.row_dimensions[4].height = 24

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 4. Filas de datos
    ofertas_ordenadas = sorted(ofertas, key=lambda p: getattr(p, "total", p.get("total", 0) if isinstance(p, dict) else 0))

    best_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde claro
    worst_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Rojo claro

    for idx, prod in enumerate(ofertas_ordenadas, start=1):
        row_num = idx + 4
        
        # Soportar objetos y dicts
        tienda = getattr(prod, "tienda", prod.get("tienda") if isinstance(prod, dict) else "")
        titulo = getattr(prod, "titulo", prod.get("titulo") if isinstance(prod, dict) else "")
        desc = getattr(prod, "descuento_porcentaje", prod.get("descuento_porcentaje", 0) if isinstance(prod, dict) else 0)
        precio = getattr(prod, "precio_actual", getattr(prod, "precio", prod.get("precio_actual", 0) if isinstance(prod, dict) else 0))
        envio = getattr(prod, "envio", prod.get("envio", 0) if isinstance(prod, dict) else 0)
        total = getattr(prod, "total", prod.get("total", 0) if isinstance(prod, dict) else 0)
        url = getattr(prod, "url", prod.get("url", "") if isinstance(prod, dict) else "")

        row_data = [
            idx,
            tienda,
            titulo,
            f"-{desc}%" if desc > 0 else "—",
            precio,
            envio,
            total,
            url
        ]
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 22

        # Formateo de celdas
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)

            if idx == 1:
                cell.fill = best_fill
            elif idx == len(ofertas_ordenadas) and len(ofertas_ordenadas) > 1:
                cell.fill = worst_fill

            if col_idx in [5, 6, 7]:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [1, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 8 and url:
                cell.hyperlink = url
                cell.value = "Abrir Oferta"
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 5. Ajustar anchos
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    wb.save(archivo_salida)
    return archivo_salida


def guardar_historial(termino: str, productos: list) -> None:
    """Guarda la búsqueda en el archivo de historial JSON."""
    if not os.path.exists(DIRECTORIO_OUTPUT):
        os.makedirs(DIRECTORIO_OUTPUT, exist_ok=True)

    historial = []
    if os.path.exists(ARCHIVO_HISTORIAL):
        try:
            with open(ARCHIVO_HISTORIAL, "r", encoding="utf-8") as f:
                historial = json.load(f)
        except Exception:
            historial = []

    registro = {
        "timestamp": datetime.now().isoformat(),
        "termino": termino,
        "total_resultados": len(productos),
        "productos": [p.to_dict() if hasattr(p, "to_dict") else p for p in productos]
    }

    historial.append(registro)

    with open(ARCHIVO_HISTORIAL, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)
